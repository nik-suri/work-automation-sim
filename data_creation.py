import sys
import pandas as pd
from openpyxl import load_workbook, Workbook
from config import *
from util import *


def main():
    args = read_command( sys.argv[1:] )

    if args.clean_prob: clean_prob()
    if args.clean_proj: clean_proj()
    if args.clean_emp: clean_emp()

    print('finished data cleaning with params: ' + str(args))


def clean_prob():
    # clean automation susceptibility datasheet
    auto_sus_df = pd.read_excel(RAW_FILES_LOC + 'frey_osborne_automation_susceptibility.xlsx')
    auto_sus_df = auto_sus_df[['Probability', 'SOC code']]
    auto_sus_df = auto_sus_df.rename(columns={'SOC code': 'SOC_CODE', 'Probability': 'AUTO_PROB'})
    auto_sus_df.to_excel(CLEAN_FILES_LOC + 'automation_susceptibility.xlsx', index=False)

    print('automation probabilities written to automation_susceptibility.xlsx')


def clean_proj():
    for msa in CA_MSA_MAP.keys():
        proj_files = CA_MSA_MAP[msa]
        clean_proj_files(proj_files)
        aggregate_proj(proj_files, msa)
        print('Cleaned and aggregated projection data for ' + msa + ' in ' + msa + '.xlsx')
        print(OUTPUT_SEPARATOR)

    print('CA employment projections cleaned and aggregated in ' + CLEAN_FILES_LOC + CLEAN_PROJECTIONS_LOC)


def clean_emp():
    # clean employment data
    for msa in CA_MSA_MAP.keys():
        msa_employment_df = pd.read_excel(RAW_FILES_LOC + RAW_EMPLOYMENT_LOC + 'MSA_M2018_dl.xlsx')
        msa_filtered = msa_employment_df.query('AREA_NAME == "' + msa + '" and TOT_EMP != "**"')[['OCC_CODE', 'TOT_EMP', 'OCC_TITLE']]
        msa_cleaned = msa_filtered.rename(columns={'OCC_CODE': 'SOC_CODE'})
        msa_cleaned.to_excel(CLEAN_FILES_LOC + CLEAN_EMPLOYMENT_LOC + msa + '.xlsx', index=False)
        print('Cleaned ' + msa + ' employment data in ' + msa + '.xlsx')
        print(OUTPUT_SEPARATOR)

    print('CA employment data extracted, cleaned, and written in ' + CLEAN_FILES_LOC + CLEAN_EMPLOYMENT_LOC)


# aggregate granular county projections into a mean metropolitan statistical area (msa) projection
def aggregate_proj(proj_files, out_filename):
    aggregate_df = pd.read_excel(CLEAN_FILES_LOC + CLEAN_PROJECTIONS_LOC + CLEANED_PREFIX + proj_files[0])

    for i in range(1, len(proj_files)):
        to_add_df = pd.read_excel(CLEAN_FILES_LOC + CLEAN_PROJECTIONS_LOC + CLEANED_PREFIX + proj_files[i])
        aggregate_df = aggregate_df.merge(to_add_df, on='SOC_CODE')
        aggregate_df['ANNUAL_CHANGE'] = aggregate_df['ANNUAL_CHANGE_x'] + aggregate_df['ANNUAL_CHANGE_y']
        del aggregate_df['ANNUAL_CHANGE_x']
        del aggregate_df['ANNUAL_CHANGE_y']

    aggregate_df['ANNUAL_MEAN_CHANGE'] = aggregate_df['ANNUAL_CHANGE'].div(len(proj_files))
    del aggregate_df['ANNUAL_CHANGE']

    aggregate_df.to_excel(CLEAN_FILES_LOC + CLEAN_PROJECTIONS_LOC + out_filename + '.xlsx', index=False)
    print('Aggregated ' + str(proj_files))


# clean California MSA local projection data excel files
def clean_proj_files(proj_files):
    for filename in proj_files:
        raw_filename = RAW_FILES_LOC + RAW_PROJECTIONS_LOC + filename
        clean_filename = CLEAN_FILES_LOC + CLEAN_PROJECTIONS_LOC + CLEANED_PREFIX + filename

        raw_wb = load_workbook(filename=raw_filename)
        clean_wb = Workbook()

        raw_ws = raw_wb.active
        clean_ws = clean_wb.active
        clean_ws.title = 'Occupational_Cleaned'
        for i in range(4, raw_ws.max_row):
            for j in range(1, raw_ws.max_column):
                clean_ws.cell(row=i - 3,column=j).value = raw_ws.cell(row=i,column=j).value

        clean_ws.delete_cols(1, 1)
        clean_ws.delete_cols(2, 4)
        clean_ws.delete_cols(3, 8)
        clean_ws.delete_rows(2, 1)

        # should delete all rows which are meaningless at the end
        empty_rows = 18
        clean_ws.delete_rows(clean_ws.max_row - empty_rows, empty_rows + 1)

        for row, cellObj in enumerate(list(clean_ws.columns)[1]):
            if row != 0:
                cellObj.value = (float(cellObj.value) + 1) ** 0.1 - 1

        clean_ws['A1'] = 'SOC_CODE'
        clean_ws['B1'] = 'ANNUAL_CHANGE'

        clean_wb.save(filename=clean_filename)
        print('Cleaned ' + filename)


def read_command(argv):
    """
    Read command line options.
    """
    from optparse import OptionParser

    parser = OptionParser()

    parser.add_option('--clean-prob', dest='clean_prob',
                      default=False, action='store_true',
                      help='clean automation probabilities excel sheet')
    parser.add_option('--clean-proj', dest='clean_proj',
                      default=False, action='store_true',
                      help='clean and aggregate regional projections')
    parser.add_option('--clean-emp', dest='clean_emp',
                      default=False, action='store_true',
                      help='clean metropolitan employment data')

    options, otherjunk = parser.parse_args(argv)
    if len(otherjunk) != 0:
        raise Exception('Command line input not understood: ' + str(otherjunk))

    return options


if __name__ == "__main__":
    main()
