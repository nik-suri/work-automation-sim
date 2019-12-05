import sys
import pandas as pd
from progress.bar import Bar
from openpyxl import load_workbook, Workbook
from config import *
from util import *


def main():
    args = read_command()
    print_warning('Cleaning data with params: ' + str(vars(args)))

    print(OUTPUT_SEPARATOR)

    if args.clean_prob: clean_prob()

    if args.run_reg:
        if args.clean_proj: clean_reg_proj()
        if args.clean_emp: clean_reg_emp()
        if args.merge: merge_reg()

    if args.run_nat:
        print_header('Cleaning national data')
        if args.clean_proj: clean_nat_proj()
        if args.clean_emp: clean_nat_emp()
        if args.merge: merge_nat()

    print_success('Finished data cleaning.')


def clean_nat_proj():
    progress_bar = Bar('Projections', max=3, suffix='%(percent)d%%')

    raw_wb = load_workbook(filename=RAW_PROJECTIONS_NAT)
    clean_wb = Workbook()

    progress_bar.next()

    raw_ws = raw_wb['Table 1.2']
    clean_ws = clean_wb.active
    clean_ws.title = 'Projections_Cleaned'
    for i in range(4, raw_ws.max_row):
        for j in range(1, raw_ws.max_column):
            clean_ws.cell(row=i - 3,column=j).value = raw_ws.cell(row=i,column=j).value

    progress_bar.next()

    clean_ws.delete_cols(1, 1)
    clean_ws.delete_cols(2, 6)

    for row, cellObj in enumerate(list(clean_ws.columns)[1]):
        if row != 0:
            cellObj.value = (float(cellObj.value) / 100 + 1) ** 0.1 - 1

    progress_bar.next()

    clean_ws['A1'] = 'SOC_CODE'
    clean_ws['B1'] = 'ANNUAL_CHANGE'

    clean_wb.save(filename=CLEAN_PROJECTIONS_NAT)
    progress_bar.finish()


def clean_nat_emp():
    progress_bar = Bar('Employment', max=3, suffix='%(percent)d%%')

    # clean employment data
    msa_employment_df = pd.read_excel(RAW_EMPLOYMENT_NAT)
    progress_bar.next()

    msa_filtered = msa_employment_df[['OCC_CODE', 'TOT_EMP', 'OCC_TITLE']]
    progress_bar.next()

    msa_cleaned = msa_filtered.rename(columns={'OCC_CODE': 'SOC_CODE'})
    progress_bar.next()

    msa_cleaned.to_excel(CLEAN_EMPLOYMENT_NAT, index=False)
    progress_bar.finish()


def merge_nat():
    progress_bar = Bar('Merging data', max=5, suffix='%(percent)d%%')

    auto_susceptibility_df = pd.read_excel(CLEAN_FREY_OSBORNE)
    progress_bar.next()

    employment_df = pd.read_excel(CLEAN_EMPLOYMENT_NAT)
    progress_bar.next()

    employment_proj_df = pd.read_excel(CLEAN_PROJECTIONS_NAT)
    progress_bar.next()

    full_employment_df = employment_df.merge(employment_proj_df, on='SOC_CODE')
    progress_bar.next()

    full_df = full_employment_df.merge(auto_susceptibility_df, on='SOC_CODE')
    progress_bar.next()

    full_df.to_excel(CLEAN_MERGED_NAT, index=False)
    progress_bar.finish()


def merge_reg():
    print_header('Merging MSA data sets')

    auto_susceptibility_df = pd.read_excel(CLEAN_FREY_OSBORNE)

    for msa in CA_MSA_MAP.keys():
        progress_bar = Bar(msa, max=4, suffix='%(percent)d%%')

        msa_filename = msa + '.xlsx'
        merged_filename = CLEAN_MERGED_MSA + msa_filename

        employment_df = pd.read_excel(CLEAN_EMPLOYMENT_MSA + msa_filename)
        progress_bar.next()

        employment_proj_df = pd.read_excel(CLEAN_PROJECTIONS_MSA + msa_filename)
        progress_bar.next()

        employment_full_df = employment_df.merge(employment_proj_df, on='SOC_CODE')
        progress_bar.next()

        full_df = employment_full_df.merge(auto_susceptibility_df, on='SOC_CODE')
        progress_bar.next()

        full_df.to_excel(merged_filename, index=False)
        progress_bar.finish()

    print_success('Full cleaned CA data files written to ' + CLEAN_MERGED_MSA)


def clean_prob():
    progress_bar = Bar('Cleaning occupational automation probabilities', max=3, suffix='%(percent)d%%')

    # clean automation susceptibility datasheet
    auto_sus_df = pd.read_excel(RAW_FREY_OSBORNE)
    progress_bar.next()

    auto_sus_df = auto_sus_df[['Probability', 'SOC code']]
    progress_bar.next()

    auto_sus_df = auto_sus_df.rename(columns={'SOC code': 'SOC_CODE', 'Probability': 'AUTO_PROB'})
    progress_bar.next()

    auto_sus_df.to_excel(CLEAN_FREY_OSBORNE, index=False)
    progress_bar.finish()


def clean_reg_proj():
    print_header('Cleaning and aggregating regional data sheets')

    for msa in CA_MSA_MAP.keys():
        proj_files = CA_MSA_MAP[msa]
        progress_bar = Bar(', '.join(proj_files), max=2*len(proj_files), suffix='%(percent)d%%')
        clean_reg_(proj_files, progress_bar)
        aggregate_reg_(proj_files, msa, progress_bar)
        progress_bar.finish()

    print_success('CA employment projections cleaned and aggregated in ' + CLEAN_PROJECTIONS)
    print(OUTPUT_SEPARATOR)


def clean_reg_emp():
    print_header('Cleaning MSA employment data')

    # clean employment data
    msa_employment_df = pd.read_excel(RAW_EMPLOYMENT_MSA)
    for msa in CA_MSA_MAP.keys():
        progress_bar = Bar(msa, max=4, suffix='%(percent)d%%')

        clean_filename = CLEAN_EMPLOYMENT_MSA + msa + '.xlsx'
        progress_bar.next()

        msa_queried = msa_employment_df.query('AREA_NAME == "' + msa + '" and TOT_EMP != "**"')
        progress_bar.next()

        msa_filtered = msa_queried[['OCC_CODE', 'TOT_EMP', 'OCC_TITLE']]
        progress_bar.next()

        msa_cleaned = msa_filtered.rename(columns={'OCC_CODE': 'SOC_CODE'})
        progress_bar.next()

        msa_cleaned.to_excel(clean_filename, index=False)
        progress_bar.finish()

    print_success('CA employment data extracted, cleaned, and written in ' + CLEAN_EMPLOYMENT_MSA)
    print(OUTPUT_SEPARATOR)


# aggregate granular county projections into a mean metropolitan statistical area (msa) projection
def aggregate_reg_(proj_files, out_filename, progress_bar):
    aggregated_filename = CLEAN_PROJECTIONS_MSA + out_filename + '.xlsx'

    aggregate_df = pd.read_excel(CLEAN_PROJECTIONS_REGIONAL + proj_files[0])
    progress_bar.next()

    for i in range(1, len(proj_files)):
        to_add_df = pd.read_excel(CLEAN_PROJECTIONS_REGIONAL + proj_files[i])
        aggregate_df = aggregate_df.merge(to_add_df, on='SOC_CODE')
        aggregate_df['ANNUAL_CHANGE'] = aggregate_df['ANNUAL_CHANGE_x'] + aggregate_df['ANNUAL_CHANGE_y']
        del aggregate_df['ANNUAL_CHANGE_x']
        del aggregate_df['ANNUAL_CHANGE_y']

        progress_bar.next()

    aggregate_df['ANNUAL_MEAN_CHANGE'] = aggregate_df['ANNUAL_CHANGE'].div(len(proj_files))
    del aggregate_df['ANNUAL_CHANGE']

    aggregate_df.to_excel(aggregated_filename, index=False)


# clean California MSA local projection data excel files
def clean_reg_(proj_files, progress_bar):
    for filename in proj_files:
        raw_filename = RAW_PROJECTIONS_REGIONAL + filename
        clean_filename = CLEAN_PROJECTIONS_REGIONAL + filename

        raw_wb = load_workbook(filename=raw_filename)
        clean_wb = Workbook()

        raw_ws = raw_wb.active
        clean_ws = clean_wb.active
        clean_ws.title = 'Occupational_Cleaned'
        for i in range(4, raw_ws.max_row):
            for j in range(1, raw_ws.max_column):
                clean_ws.cell(row=i - 3,column=j).value = raw_ws.cell(row=i,column=j).value

        clean_ws.delete_cols(1, 1)
        clean_ws.delete_cols(2, 4)
        clean_ws.delete_cols(3, 8)
        clean_ws.delete_rows(2, 1)

        # should delete all rows which are meaningless at the end
        empty_rows = 18
        clean_ws.delete_rows(clean_ws.max_row - empty_rows, empty_rows + 1)

        for row, cellObj in enumerate(list(clean_ws.columns)[1]):
            if row != 0:
                cellObj.value = (float(cellObj.value) + 1) ** 0.1 - 1

        clean_ws['A1'] = 'SOC_CODE'
        clean_ws['B1'] = 'ANNUAL_CHANGE'

        clean_wb.save(filename=clean_filename)
        progress_bar.next()


def read_command():
    """
    Read command line arguments.
    """
    from argparse import ArgumentParser

    parser = ArgumentParser(description=('Clean data for the simulation.'
                                         'Default behavior is to clean and merge all national-level and regional-level data files.'
                                         'Use options to run only specified steps of the program.'))

    parser.add_argument('--clean-prob', dest='clean_prob',
                        default=False, action='store_true',
                        help='clean automation probabilities excel sheet')
    parser.add_argument('--clean-proj', dest='clean_proj',
                        default=False, action='store_true',
                        help='clean and aggregate regional projections')
    parser.add_argument('--clean-emp', dest='clean_emp',
                        default=False, action='store_true',
                        help='clean metropolitan employment data')
    parser.add_argument('--merge', dest='merge',
                        default=False, action='store_true',
                        help='merge files into full MSA dataframes including probability, employment, and projections')
    parser.add_argument('--nat', dest='run_nat',
                        default=False, action='store_true',
                        help='clean only national-level data files')
    parser.add_argument('--reg', dest='run_reg',
                        default=False, action='store_true',
                        help='clean only regional-level data files')

    args = parser.parse_args()

    if all([not args.clean_prob, not args.clean_proj, not args.clean_emp, not args.merge]):
        print_warning('No cleaning options specified. Running with options set to clean all data.')
        args.clean_prob = True
        args.clean_proj = True
        args.clean_emp = True
        args.merge = True

    if all([not args.run_nat, not args.run_reg]):
        print_warning('National/Regional-level cleaning not specified. Running with options set to clean both levels of data.')
        args.run_nat = True
        args.run_reg = True

    return args


if __name__ == "__main__":
    main()
